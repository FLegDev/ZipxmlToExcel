# invoices/admin.py
from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.admin import DateFieldListFilter
from django.http import HttpResponse
from django.urls import reverse
from django.utils.html import format_html
from django.utils.translation import gettext as _
from django.utils.translation import gettext as __

from .models import Invoice, InvoiceItem, InvoiceUpload, UserProfile
from .services import import_zip_file

import openpyxl
from datetime import datetime

# ============ ACTION EXCEL =============

@admin.action(description=_("Exporter (feuilles séparées)"))
def export_invoices_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Factures"

    headers_invoice = [
        __("N°"),
        __("Série"),
        __("Date"),
        __("Acheteur"),
        __("SIREN acheteur"),
        __("Total HT"),
        __("TVA"),
        __("Total TTC"),
    ]
    ws1.append(headers_invoice)

    for invoice in queryset:
        ws1.append([
            invoice.invoice_number,
            invoice.invoice_symbol,
            invoice.issued_date.strftime('%m/%d/%Y') if invoice.issued_date else "",
            invoice.buyer_name,
            invoice.buyer_tax_code,
            float(invoice.total_before_tax),
            float(invoice.total_tax),
            float(invoice.total_after_tax),
        ])

    ws2 = wb.create_sheet(title="Lignes de facture")

    headers_items = [
        __("Facture"),
        __("N° ligne"),
        __("Désignation"),
        __("Unité"),
        __("Quantité"),
        __("Prix unitaire"),
        __("Montant HT"),
        __("TVA"),
    ]
    ws2.append(headers_items)

    for invoice in queryset:
        for item in invoice.items.all():
            ws2.append([
                f"{invoice.invoice_symbol}-{invoice.invoice_number}",
                item.line_number,
                item.name,
                item.unit,
                item.quantity,
                float(item.unit_price),
                float(item.total_price),
                item.vat_rate if item.vat_rate else ""
            ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"factures_export_{datetime.now().strftime('%m/%d/%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@admin.action(description=_("Exporter fusionné (modèle structuré)"))
def export_invoices_fusionned(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures détaillées"

    headers = [
        __("STT"),
        __("Tên"),
        __("DVT"),
        __("SL"),
        __("Đơn giá"),
        __("Thuế suất"),
        __("Thành tiền"),
        __("Số HĐ"),
        __("MST"),
        __("Tên cty"),
        __("Ngày"),
    ]
    ws.append(headers)

    for invoice in queryset:
        for item in invoice.items.all():
            ws.append([
                item.line_number,
                item.name,
                item.unit,
                item.quantity,
                float(item.unit_price),
                item.vat_rate or "",
                float(item.total_price),
                invoice.invoice_number,
                invoice.buyer_tax_code,
                invoice.buyer_name,
                invoice.issued_date.strftime('%m/%d/%Y') if invoice.issued_date else ""
            ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"factures_fusion_structuree_{datetime.now().strftime('%m/%d/%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ============ INVOICE & ITEMS =============

class InvoiceItemInline(admin.TabularInline):
    model = InvoiceItem
    extra = 0
    readonly_fields = ['line_number', 'name', 'unit', 'quantity', 'unit_price', 'total_price', 'vat_rate']

@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    change_list_template = "admin/invoices/invoice/change_list.html"
    list_display = (
        'invoice_symbol', 'invoice_number', 'issued_date',
        'buyer_name', 'buyer_tax_code', 'total_after_tax', 'import_date'
    )
    list_filter = (
        'currency',
        ('issued_date', DateFieldListFilter),
        ('import_date', DateFieldListFilter),
    )
    search_fields = ('invoice_number', 'buyer_name', 'buyer_tax_code')
    inlines = [InvoiceItemInline]
    actions = [export_invoices_to_excel, export_invoices_fusionned]

    def import_zip_link(self, obj):
        url = reverse('upload_multiple_zips')  # name de la vue
        return format_html('<a class="button" href="{}">📥 Importer des ZIPs</a>', url)
    import_zip_link.short_description = "Import ZIP"

# ============ ZIP IMPORT =============

@admin.register(InvoiceUpload)
class InvoiceUploadAdmin(admin.ModelAdmin):
    list_display = ('file', 'uploaded_at')

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if obj.file.name.endswith('.zip'):
            import_zip_file(obj.file)

# ============ PROFIL UTILISATEUR =============

@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'preferred_language')
    list_filter = ('preferred_language',)

class UserProfileInline(admin.StackedInline):
    model = UserProfile
    can_delete = False
    verbose_name_plural = "Profil utilisateur"
    fk_name = "user"

class UserAdmin(BaseUserAdmin):
    inlines = (UserProfileInline,)

    def get_inline_instances(self, request, obj=None):
        if not obj:
            return []
        return super().get_inline_instances(request, obj)

admin.site.unregister(User)
admin.site.register(User, UserAdmin)
